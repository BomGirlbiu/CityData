# 爬取B站视频下的评论
import random
import requests
import re
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import logging
import os

# 创建log文件夹
if not os.path.exists("./log"):
    os.makedirs("./log")
if not os.path.exists("./temp"):
    os.makedirs("./temp")
# 配置日志
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler(
                            "./log/get_comments.log", encoding='utf-8'),
                        logging.StreamHandler()
                    ])

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}


def get_video_id(bv):
    url = f'https://www.bilibili.com/video/{bv}'
    html = requests.get(url, headers=headers)
    html.encoding = 'utf-8'
    content = html.text
    aid_regx = '"aid":(.*?),"bvid":"{}"'.format(bv)
    video_aid = re.findall(aid_regx, content)[0]
    return video_aid


def fetch_comment_replies(video_id, comment_id, parent_user_name, max_pages=1):
    replies = []
    preLen = 0
    for page in range(1, max_pages + 1):
        url = f'https://api.bilibili.com/x/v2/reply/reply?oid={
            video_id}&type=1&root={comment_id}&ps=10&pn={page}'
        try:
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 200:
                logging.info(f'  in page {page} - good')
                data = response.json()
                if data and data.get('data') and 'replies' in data['data']:
                    for reply in data['data']['replies']:
                        reply_info = {
                            '用户昵称': reply['member']['uname'],
                            '评论内容': reply['content']['message'],
                            '评论层级': '二级评论',
                            '点赞数量': reply['like'],
                            '回复时间': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(reply['ctime']))
                        }
                        replies.append(reply_info)
                    if preLen == len(replies):
                        break
                    preLen = len(replies)
                else:
                    return replies
            else:
                logging.info(f'  in page {page} - bad')
        except requests.RequestException as e:
            logging.error(f"请求出错: {e}")
            break
        # 控制请求频率，1到2秒之间
        time.sleep(random.uniform(1, 2))
    return replies


def fetch_comments(video_id, max_pages=3):
    comments = []
    last_count = 0
    for page in range(1, max_pages + 1):
        url = f'https://api.bilibili.com/x/v2/reply?pn={
            page}&type=1&oid={video_id}&sort=2'
        try:
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 200:
                logging.info(f'in page {page} - good')
                data = response.json()
                if data and 'replies' in data['data']:
                    cnt = 0
                    if not data['data']['replies']:
                        continue
                    for comment in data['data']['replies']:
                        comment_info = {
                            '用户昵称': comment['member']['uname'],
                            '评论内容': comment['content']['message'],
                            '评论层级': '一级评论',
                            '点赞数量': comment['like'],
                            '回复时间': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(comment['ctime']))
                        }
                        comments.append(comment_info)
                        if cnt < 3:
                            replies = fetch_comment_replies(
                                video_id, comment['rpid'], comment['member']['uname'])
                            comments.extend(replies)
                            cnt += 1
                if last_count == len(comments):
                    break
                last_count = len(comments)
            else:
                logging.info(f'in page {page} - bad')
                break
        except requests.RequestException as e:
            logging.error(f"请求出错: {e}")
            break
        except KeyError:
            logging.error("Expected key missing in the response data")
            break
        # 控制请求频率，1到2秒之间
        time.sleep(random.uniform(1, 2))
    return comments


def save_comments_to_xlsx(city, comments, video_bv):
    file_path = f'./temp/{city}_comments.xlsx'

    try:
        # 尝试加载现有的工作簿
        wb = load_workbook(file_path)
        ws = wb.active
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = video_bv
        headers = ['城市', '视频bv', '用户昵称', '评论内容', '评论层级', '点赞数量', '回复时间']
        ws.append(headers)

    # 找到工作表的最后一行并追加数据
    for comment in comments:
        ws.append([
            city,
            video_bv,
            comment['用户昵称'],
            comment['评论内容'],
            comment['评论层级'],
            comment['点赞数量'],
            comment['回复时间']
        ])

    wb.save(file_path)


filename = './temp/target/combined_news_videos.xlsx'

# 打开文件并读取数据
wb = openpyxl.load_workbook(filename)
ws = wb.active

cnt = 0
lastname = 't'
fg = False

for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过标题行
    city = row[0]  # 城市名

    # if city == '松原':
    #     fg = True
    # if not fg:
    #     continue

    video_name = row[3]  # 视频名字
    video_bv = row[2]  # video_bv
    if video_bv.find('BV') == -1:
        continue
    video_bv = re.search(r'/video/(\w+)/', video_bv).group(1)
    # 计数
    if lastname != city:
        cnt = 0
        lastname = city
    cnt += 1
    logging.info('(%d)城市: %s, 视频名字: %s, video_bv: %s',
                 cnt, city, video_name, video_bv)
    aid = get_video_id(video_bv)
    video_id = aid
    comments = fetch_comments(video_id)
    save_comments_to_xlsx(city, comments, video_bv)

logging.info('所有评论已成功保存到相应的Excel文件中。')
